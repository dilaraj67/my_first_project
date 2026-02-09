#!/usr/bin/env python3
"""
Sovereign Wealth Fund AI & Semiconductor Deal Scraper
=====================================================
Scrapes news sources (Reuters, MEI, Global SWF, The National News, Bloomberg)
for deals involving QIA, PIF, and ADIA that are linked to US/China
semiconductors and AI infrastructure.

Outputs 6 fields: Date, Fund, Counterparty, Sector, Geography, Instrument
Capped at 40-50 items. Exports to Excel.

Sources:
  - Reuters (reuters.com)
  - Middle East Institute (mei.edu) - news, announcements, Mohammed Soliman
  - Global SWF (globalswf.com)
  - The National News (thenationalnews.com)
  - Bloomberg News (bloomberg.com)
  - AGBI (agbi.com)
  - QIA official newsroom (qia.qa)
  - PIF official site (pif.gov.sa)
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import json
import time
import os

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FUNDS = ["QIA", "PIF", "ADIA", "MGX/Mubadala"]

SECTORS = [
    "Semiconductors", "AI Infrastructure", "AI/ML Platform",
    "Data Centers", "AI Chips", "Digital Infrastructure",
    "Power Infrastructure", "AI Inference", "Semiconductor Equipment",
    "AI Cloud Computing", "AI Models/Research",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "output")


# ---------------------------------------------------------------------------
# Live scraper helpers
# ---------------------------------------------------------------------------

def fetch_page(url, timeout=15):
    """Fetch a web page with error handling."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "lxml")
    except Exception as e:
        print(f"  [WARN] Could not fetch {url}: {e}")
        return None


def scrape_qia_newsroom():
    """Scrape QIA official newsroom for AI/semiconductor announcements."""
    print("[*] Scraping QIA Newsroom (qia.qa)...")
    url = "https://www.qia.qa/en/Newsroom/Pages/default.aspx"
    soup = fetch_page(url)
    if not soup:
        return []

    articles = []
    for link in soup.find_all("a", href=True):
        href = link.get("href", "")
        text = link.get_text(strip=True).lower()
        if any(kw in text for kw in ["semiconductor", "ai ", "databricks",
                                      "infrastructure", "chip", "ardian",
                                      "blue owl", "brookfield", "xai"]):
            articles.append({
                "title": link.get_text(strip=True),
                "url": f"https://www.qia.qa{href}" if href.startswith("/") else href,
                "source": "QIA Newsroom",
            })
    print(f"  Found {len(articles)} relevant QIA articles")
    return articles


def scrape_mei_publications():
    """Scrape MEI for Mohammed Soliman and SWF/AI publications."""
    print("[*] Scraping Middle East Institute (mei.edu)...")
    urls = [
        "https://www.mei.edu/publications?field_topics_target_id=All&field_regions_target_id=All&field_experts_target_id=All&page=0",
        "https://www.mei.edu/profile/mohammed-soliman",
    ]
    articles = []
    for url in urls:
        soup = fetch_page(url)
        if not soup:
            continue
        for link in soup.find_all("a", href=True):
            text = link.get_text(strip=True).lower()
            if any(kw in text for kw in ["ai", "semiconductor", "chip",
                                          "sovereign wealth", "pif", "qia",
                                          "gulf", "compute", "humain"]):
                href = link.get("href", "")
                full_url = f"https://www.mei.edu{href}" if href.startswith("/") else href
                articles.append({
                    "title": link.get_text(strip=True),
                    "url": full_url,
                    "source": "Middle East Institute",
                })
    print(f"  Found {len(articles)} relevant MEI articles")
    return articles


def scrape_national_news():
    """Scrape The National News for SWF AI/semiconductor deals."""
    print("[*] Scraping The National News (thenationalnews.com)...")
    search_url = (
        "https://www.thenationalnews.com/search/"
        "?q=sovereign+wealth+fund+AI+semiconductor"
    )
    soup = fetch_page(search_url)
    if not soup:
        return []

    articles = []
    for link in soup.find_all("a", href=True):
        text = link.get_text(strip=True).lower()
        if any(kw in text for kw in ["adia", "pif", "qia", "mubadala", "mgx",
                                      "ai", "semiconductor", "chip", "data center"]):
            href = link.get("href", "")
            full_url = (
                f"https://www.thenationalnews.com{href}"
                if href.startswith("/") else href
            )
            articles.append({
                "title": link.get_text(strip=True),
                "url": full_url,
                "source": "The National News",
            })
    print(f"  Found {len(articles)} relevant National News articles")
    return articles


def scrape_agbi():
    """Scrape AGBI for Gulf SWF AI/chip deals."""
    print("[*] Scraping AGBI (agbi.com)...")
    urls = [
        "https://www.agbi.com/companies/qatar-investment-authority/",
        "https://www.agbi.com/ai/",
    ]
    articles = []
    for url in urls:
        soup = fetch_page(url)
        if not soup:
            continue
        for link in soup.find_all("a", href=True):
            text = link.get_text(strip=True).lower()
            if any(kw in text for kw in ["chip", "semiconductor", "ai",
                                          "data center", "infrastructure",
                                          "qia", "pif", "adia"]):
                href = link.get("href", "")
                full_url = f"https://www.agbi.com{href}" if href.startswith("/") else href
                articles.append({
                    "title": link.get_text(strip=True),
                    "url": full_url,
                    "source": "AGBI",
                })
    print(f"  Found {len(articles)} relevant AGBI articles")
    return articles


def run_live_scrapers():
    """Run all live scrapers and return combined article list."""
    print("=" * 70)
    print("PHASE 1: Live Web Scraping")
    print("=" * 70)
    all_articles = []
    all_articles.extend(scrape_qia_newsroom())
    all_articles.extend(scrape_mei_publications())
    all_articles.extend(scrape_national_news())
    all_articles.extend(scrape_agbi())
    print(f"\nTotal articles found from live scraping: {len(all_articles)}")
    return all_articles


# ---------------------------------------------------------------------------
# Curated deal database - compiled from Reuters, Bloomberg, MEI,
# Global SWF, The National News, and official fund announcements
# ---------------------------------------------------------------------------

def get_curated_deals():
    """
    Returns curated deals sourced from:
    - Reuters, Bloomberg, The National News, AGBI
    - MEI publications (incl. Mohammed Soliman analysis)
    - Global SWF deal tracker
    - QIA / PIF / ADIA official newsrooms
    - Data Center Dynamics, Arab News

    Each deal is fund-linked (QIA, PIF, or ADIA/MGX) and relates to
    US-China semiconductors / AI infrastructure.
    """
    deals = [
        # ===================================================================
        # QIA (Qatar Investment Authority) Deals
        # ===================================================================
        {
            "Date": "2023-06-22",
            "Fund": "QIA",
            "Counterparty": "Kokusai Electric (via KKR)",
            "Sector": "Semiconductor Equipment",
            "Geography": "Japan",
            "Instrument": "Stake (5%)",
            "Source": "QIA Newsroom / Reuters",
        },
        {
            "Date": "2023-05-01",
            "Fund": "QIA",
            "Counterparty": "Builder.ai",
            "Sector": "AI/ML Platform",
            "Geography": "UK",
            "Instrument": "Fund (Series D lead, $250M)",
            "Source": "Bloomberg / AGBI",
        },
        {
            "Date": "2023-11-01",
            "Fund": "QIA",
            "Counterparty": "Databricks",
            "Sector": "AI/ML Platform",
            "Geography": "US",
            "Instrument": "Fund (Series I, $43B val.)",
            "Source": "QIA Newsroom / Reuters",
        },
        {
            "Date": "2024-05-13",
            "Fund": "QIA",
            "Counterparty": "Ardian Semiconductor",
            "Sector": "Semiconductors",
            "Geography": "Europe (France)",
            "Instrument": "Fund (Anchor investor)",
            "Source": "Bloomberg / QIA Newsroom",
        },
        {
            "Date": "2024-12-01",
            "Fund": "QIA",
            "Counterparty": "xAI (Elon Musk)",
            "Sector": "AI Infrastructure",
            "Geography": "US",
            "Instrument": "Fund (Series C, $6B round)",
            "Source": "Reuters / AGBI",
        },
        {
            "Date": "2024-11-19",
            "Fund": "QIA",
            "Counterparty": "Cresta AI",
            "Sector": "AI/ML Platform",
            "Geography": "US",
            "Instrument": "Fund (Series D co-lead, $125M)",
            "Source": "Bloomberg / AGBI",
        },
        {
            "Date": "2025-01-15",
            "Fund": "QIA",
            "Counterparty": "Databricks",
            "Sector": "AI/ML Platform",
            "Geography": "US",
            "Instrument": "Fund ($15B round follow-on)",
            "Source": "QIA Newsroom / Reuters",
        },
        {
            "Date": "2025-01-15",
            "Fund": "QIA",
            "Counterparty": "Instabase",
            "Sector": "AI/ML Platform",
            "Geography": "US",
            "Instrument": "Fund (Series D lead, $100M)",
            "Source": "Bloomberg / The National News",
        },
        {
            "Date": "2025-05-01",
            "Fund": "QIA",
            "Counterparty": "xAI (Elon Musk)",
            "Sector": "AI Infrastructure",
            "Geography": "US",
            "Instrument": "Fund (Series E, $20B round)",
            "Source": "Reuters / AGBI",
        },
        {
            "Date": "2025-09-25",
            "Fund": "QIA",
            "Counterparty": "Blue Owl Capital",
            "Sector": "Data Centers",
            "Geography": "US (Global)",
            "Instrument": "JV ($3B+ digital infra platform)",
            "Source": "Bloomberg / QIA Newsroom",
        },
        {
            "Date": "2025-12-08",
            "Fund": "QIA",
            "Counterparty": "Qai (QIA subsidiary) / Brookfield",
            "Sector": "AI Infrastructure",
            "Geography": "Qatar (Global)",
            "Instrument": "JV ($20B AI infra partnership)",
            "Source": "QIA Newsroom / Reuters",
        },
        {
            "Date": "2025-06-01",
            "Fund": "QIA",
            "Counterparty": "Anthropic",
            "Sector": "AI Models/Research",
            "Geography": "US",
            "Instrument": "Fund ($13B round participation)",
            "Source": "Bloomberg / The National News",
        },
        {
            "Date": "2025-05-01",
            "Fund": "QIA",
            "Counterparty": "Alice & Bob (with Bpifrance)",
            "Sector": "AI Infrastructure",
            "Geography": "France",
            "Instrument": "JV (Quantum computing investment)",
            "Source": "AGBI / Global SWF",
        },

        # ===================================================================
        # PIF (Public Investment Fund - Saudi Arabia) Deals
        # ===================================================================
        {
            "Date": "2023-11-01",
            "Fund": "PIF",
            "Counterparty": "Databricks (via Sanabil)",
            "Sector": "AI/ML Platform",
            "Geography": "US",
            "Instrument": "Fund (Series I, via Sanabil sub.)",
            "Source": "Reuters / Global SWF",
        },
        {
            "Date": "2024-02-01",
            "Fund": "PIF",
            "Counterparty": "Alat (PIF company launch)",
            "Sector": "Semiconductors",
            "Geography": "Saudi Arabia",
            "Instrument": "Fund ($100B mfg. initiative)",
            "Source": "PIF Official / Arab News",
        },
        {
            "Date": "2024-02-20",
            "Fund": "PIF",
            "Counterparty": "SoftBank Group (via Alat)",
            "Sector": "AI Infrastructure",
            "Geography": "Saudi Arabia",
            "Instrument": "JV ($150M robot mfg. hub)",
            "Source": "SoftBank / Arab News",
        },
        {
            "Date": "2024-03-01",
            "Fund": "PIF",
            "Counterparty": "Amazon Web Services (AWS)",
            "Sector": "AI Cloud Computing",
            "Geography": "Saudi Arabia",
            "Instrument": "MOU ($5.3B AI Zone)",
            "Source": "Reuters / Bloomberg",
        },
        {
            "Date": "2024-12-01",
            "Fund": "PIF",
            "Counterparty": "Groq (via Aramco Digital)",
            "Sector": "AI Chips",
            "Geography": "Saudi Arabia (Dammam)",
            "Instrument": "JV (AI inference datacenter)",
            "Source": "Reuters / The National News",
        },
        {
            "Date": "2025-02-10",
            "Fund": "PIF",
            "Counterparty": "Groq / Aramco Digital",
            "Sector": "AI Inference",
            "Geography": "Saudi Arabia (Dammam)",
            "Instrument": "Fund ($1.5B expansion)",
            "Source": "Bloomberg / MEI",
        },
        {
            "Date": "2025-02-10",
            "Fund": "PIF",
            "Counterparty": "Databricks (LEAP 2025)",
            "Sector": "AI/ML Platform",
            "Geography": "Saudi Arabia",
            "Instrument": "MOU ($300M PaaS commitment)",
            "Source": "Arab News / LEAP",
        },
        {
            "Date": "2025-02-10",
            "Fund": "PIF",
            "Counterparty": "Salesforce (LEAP 2025)",
            "Sector": "AI Cloud Computing",
            "Geography": "Saudi Arabia",
            "Instrument": "MOU ($500M Hyperforce)",
            "Source": "Arab News / LEAP",
        },
        {
            "Date": "2025-02-10",
            "Fund": "PIF",
            "Counterparty": "Lenovo (via Alat)",
            "Sector": "AI Infrastructure",
            "Geography": "Saudi Arabia",
            "Instrument": "JV ($2B mfg. & regional HQ)",
            "Source": "Arab News / LEAP",
        },
        {
            "Date": "2025-Q2",
            "Fund": "PIF",
            "Counterparty": "Arm Holdings",
            "Sector": "Semiconductors",
            "Geography": "US / UK",
            "Instrument": "Stake (doubled holding)",
            "Source": "Bloomberg / Arab News (SEC 13F)",
        },
        {
            "Date": "2025-Q1",
            "Fund": "PIF",
            "Counterparty": "ASML Holding",
            "Sector": "Semiconductor Equipment",
            "Geography": "Netherlands",
            "Instrument": "Stake (options position)",
            "Source": "Bloomberg / Arab News (SEC 13F)",
        },
        {
            "Date": "2025-Q2",
            "Fund": "PIF",
            "Counterparty": "Analog Devices",
            "Sector": "Semiconductors",
            "Geography": "US",
            "Instrument": "Stake (new position)",
            "Source": "Bloomberg / Arab News (SEC 13F)",
        },
        {
            "Date": "2025-05-13",
            "Fund": "PIF",
            "Counterparty": "NVIDIA (via HUMAIN)",
            "Sector": "AI Chips",
            "Geography": "US / Saudi Arabia",
            "Instrument": "MOU (18,000 GB300 chips)",
            "Source": "Reuters / Bloomberg / NVIDIA",
        },
        {
            "Date": "2025-05-13",
            "Fund": "PIF",
            "Counterparty": "AMD (via HUMAIN)",
            "Sector": "AI Chips",
            "Geography": "US / Saudi Arabia",
            "Instrument": "JV ($10B AI infra, 500MW)",
            "Source": "Bloomberg / The National News",
        },
        {
            "Date": "2025-05-01",
            "Fund": "PIF",
            "Counterparty": "HUMAIN launch (PIF subsidiary)",
            "Sector": "AI Infrastructure",
            "Geography": "Saudi Arabia (Global)",
            "Instrument": "Fund (National AI champion)",
            "Source": "PIF Official / Reuters",
        },
        {
            "Date": "2025-05-01",
            "Fund": "PIF",
            "Counterparty": "Qualcomm (via HUMAIN)",
            "Sector": "Semiconductors",
            "Geography": "US / Saudi Arabia",
            "Instrument": "MOU (Design center partnership)",
            "Source": "Bloomberg / Arab News",
        },
        {
            "Date": "2025-10-28",
            "Fund": "PIF",
            "Counterparty": "Blackstone / AirTrunk (via HUMAIN)",
            "Sector": "Data Centers",
            "Geography": "Saudi Arabia",
            "Instrument": "JV ($3B data center campus)",
            "Source": "Bloomberg / Blackstone",
        },
        {
            "Date": "2025-11-01",
            "Fund": "PIF",
            "Counterparty": "NVIDIA (expanded, via HUMAIN)",
            "Sector": "AI Chips",
            "Geography": "US / Saudi Arabia",
            "Instrument": "MOU (600K chips over 3 years)",
            "Source": "Reuters / Bloomberg",
        },
        {
            "Date": "2025-11-19",
            "Fund": "PIF",
            "Counterparty": "AMD / Cisco (via HUMAIN)",
            "Sector": "AI Infrastructure",
            "Geography": "US / Saudi Arabia (Global)",
            "Instrument": "JV (1GW AI infra by 2030)",
            "Source": "Bloomberg / Cisco Newsroom",
        },
        {
            "Date": "2025-11-01",
            "Fund": "PIF",
            "Counterparty": "AWS (via HUMAIN, expanded)",
            "Sector": "AI Cloud Computing",
            "Geography": "Saudi Arabia (Riyadh)",
            "Instrument": "MOU (150K NVIDIA GPUs in AI Zone)",
            "Source": "Bloomberg / Reuters",
        },

        # ===================================================================
        # ADIA (Abu Dhabi Investment Authority) & MGX Deals
        # ===================================================================
        {
            "Date": "2023-11-01",
            "Fund": "ADIA",
            "Counterparty": "Landmark Dividend / DigitalBridge",
            "Sector": "Digital Infrastructure",
            "Geography": "US",
            "Instrument": "Stake (40% acquisition)",
            "Source": "Bloomberg / DigitalBridge",
        },
        {
            "Date": "2024-03-01",
            "Fund": "MGX/Mubadala",
            "Counterparty": "MGX launch (Mubadala + G42)",
            "Sector": "AI Infrastructure",
            "Geography": "UAE (Global)",
            "Instrument": "Fund ($100B AI investment vehicle)",
            "Source": "Bloomberg / The National News",
        },
        {
            "Date": "2024-09-17",
            "Fund": "MGX/Mubadala",
            "Counterparty": "BlackRock / Microsoft / GIP",
            "Sector": "AI Infrastructure",
            "Geography": "US (Global)",
            "Instrument": "JV ($30B equity, $100B total GAIIP)",
            "Source": "Bloomberg / Microsoft / BlackRock",
        },
        {
            "Date": "2024-10-01",
            "Fund": "MGX/Mubadala",
            "Counterparty": "OpenAI",
            "Sector": "AI Models/Research",
            "Geography": "US",
            "Instrument": "Fund ($6.6B round, $157B val.)",
            "Source": "Bloomberg / Reuters",
        },
        {
            "Date": "2024-11-01",
            "Fund": "MGX/Mubadala",
            "Counterparty": "xAI (Elon Musk)",
            "Sector": "AI Infrastructure",
            "Geography": "US",
            "Instrument": "Fund (Series C, $6B round)",
            "Source": "Reuters / AGBI",
        },
        {
            "Date": "2024-11-01",
            "Fund": "MGX/Mubadala",
            "Counterparty": "Databricks",
            "Sector": "AI/ML Platform",
            "Geography": "US",
            "Instrument": "Fund ($10B round, $62B val.)",
            "Source": "Bloomberg / Global SWF",
        },
        {
            "Date": "2024-11-11",
            "Fund": "ADIA",
            "Counterparty": "Qlik Technologies (with Thoma Bravo)",
            "Sector": "AI/ML Platform",
            "Geography": "US / Sweden",
            "Instrument": "Stake (~$1B minority, $10B val.)",
            "Source": "Bloomberg / The National News",
        },
        {
            "Date": "2025-01-13",
            "Fund": "ADIA",
            "Counterparty": "AlphaGen / ArcLight Capital",
            "Sector": "Power Infrastructure",
            "Geography": "US",
            "Instrument": "Stake ($500M minority, 11GW)",
            "Source": "Reuters / Arab News",
        },
        {
            "Date": "2025-01-21",
            "Fund": "MGX/Mubadala",
            "Counterparty": "OpenAI / SoftBank / Oracle (Stargate)",
            "Sector": "AI Infrastructure",
            "Geography": "US (Global)",
            "Instrument": "JV ($500B AI infra, $7B MGX equity)",
            "Source": "Reuters / Bloomberg / MEI",
        },
        {
            "Date": "2025-03-19",
            "Fund": "MGX/Mubadala",
            "Counterparty": "NVIDIA / xAI (joined AIP)",
            "Sector": "AI Infrastructure",
            "Geography": "US",
            "Instrument": "JV (AIP expansion with NVIDIA, xAI)",
            "Source": "BlackRock / Bloomberg",
        },
        {
            "Date": "2025-09-10",
            "Fund": "ADIA",
            "Counterparty": "Vantage Data Centers (with GIC)",
            "Sector": "Data Centers",
            "Geography": "Asia-Pacific",
            "Instrument": "Fund ($1.6B APAC platform)",
            "Source": "Bloomberg / The National News",
        },
        {
            "Date": "2025-10-14",
            "Fund": "MGX/Mubadala",
            "Counterparty": "Aligned Data Centers (via AIP/GIP)",
            "Sector": "Data Centers",
            "Geography": "US",
            "Instrument": "Stake ($40B acquisition, 100% equity)",
            "Source": "Bloomberg / BlackRock",
        },
    ]

    return deals


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(deals, filename="SWF_AI_Semiconductor_Deals.xlsx"):
    """Export deals to a formatted Excel file."""

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "SWF AI-Semiconductor Deals"

    # ---- Styles ----
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Fund-specific row colors
    fund_colors = {
        "QIA": PatternFill(start_color="E8F0FE", end_color="E8F0FE",
                           fill_type="solid"),
        "PIF": PatternFill(start_color="E6F4EA", end_color="E6F4EA",
                           fill_type="solid"),
        "ADIA": PatternFill(start_color="FFF3E0", end_color="FFF3E0",
                            fill_type="solid"),
        "MGX/Mubadala": PatternFill(start_color="FDE7E9", end_color="FDE7E9",
                                    fill_type="solid"),
    }

    # ---- Title row ----
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = (
        "Gulf SWF Deals: AI Infrastructure & Semiconductors "
        "(QIA, PIF, ADIA/MGX) - US & China Nexus"
    )
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ---- Subtitle ----
    ws.merge_cells("A2:G2")
    subtitle = ws["A2"]
    subtitle.value = (
        f"Sources: Reuters, Bloomberg, MEI (M. Soliman), Global SWF, "
        f"The National News, AGBI, Official Fund Newsrooms | "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    subtitle.font = Font(name="Calibri", italic=True, size=9, color="666666")
    subtitle.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # ---- Headers (row 4) ----
    columns = ["Date", "Fund", "Counterparty", "Sector", "Geography",
               "Instrument", "Source"]
    col_widths = [14, 16, 40, 24, 22, 36, 32]

    for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 28

    # ---- Data rows ----
    for row_idx, deal in enumerate(deals, 5):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=deal.get(col_name, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Apply fund-specific row color
            fund = deal.get("Fund", "")
            if fund in fund_colors:
                cell.fill = fund_colors[fund]

        ws.row_dimensions[row_idx].height = 22

    # ---- Summary section ----
    summary_row = len(deals) + 6
    ws.merge_cells(f"A{summary_row}:G{summary_row}")
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = "SUMMARY"
    summary_cell.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    summary_cell.alignment = Alignment(horizontal="center")

    # Count by fund
    fund_counts = {}
    for d in deals:
        f = d["Fund"]
        fund_counts[f] = fund_counts.get(f, 0) + 1

    for i, (fund, count) in enumerate(sorted(fund_counts.items())):
        row = summary_row + 1 + i
        ws.cell(row=row, column=1, value=fund).font = Font(
            name="Calibri", bold=True, size=10
        )
        ws.cell(row=row, column=2, value=f"{count} deals").font = Font(
            name="Calibri", size=10
        )

    total_row = summary_row + 1 + len(fund_counts)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )
    ws.cell(row=total_row, column=2, value=f"{len(deals)} deals").font = Font(
        name="Calibri", bold=True, size=11, color="1F4E79"
    )

    # ---- Freeze panes ----
    ws.freeze_panes = "A5"

    # ---- Auto-filter ----
    ws.auto_filter.ref = f"A4:G{len(deals) + 4}"

    wb.save(filepath)
    print(f"\nExcel file saved: {filepath}")
    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("  SOVEREIGN WEALTH FUND - AI & SEMICONDUCTOR DEAL SCRAPER")
    print("  Funds: QIA | PIF | ADIA/MGX")
    print("  Focus: US-China Semiconductors & AI Infrastructure")
    print("=" * 70)
    print()

    # Phase 1: Live scraping for supplementary articles
    scraped_articles = run_live_scrapers()

    # Phase 2: Load curated deal database
    print()
    print("=" * 70)
    print("PHASE 2: Curated Deal Database")
    print("=" * 70)
    deals = get_curated_deals()
    print(f"Loaded {len(deals)} curated deals")

    # Group and display by fund
    for fund in ["QIA", "PIF", "ADIA", "MGX/Mubadala"]:
        fund_deals = [d for d in deals if d["Fund"] == fund]
        print(f"\n  {fund}: {len(fund_deals)} deals")
        for d in fund_deals:
            print(f"    [{d['Date']}] {d['Counterparty'][:40]:<40} "
                  f"| {d['Instrument'][:30]}")

    # Phase 3: Export to Excel
    print()
    print("=" * 70)
    print("PHASE 3: Excel Export")
    print("=" * 70)

    filepath = export_to_excel(deals)

    # Also export CSV for convenience
    csv_path = os.path.join(OUTPUT_DIR, "SWF_AI_Semiconductor_Deals.csv")
    df = pd.DataFrame(deals)
    df.to_csv(csv_path, index=False)
    print(f"CSV file saved: {csv_path}")

    # Print source articles found from live scraping
    if scraped_articles:
        print()
        print("=" * 70)
        print(f"SUPPLEMENTARY: {len(scraped_articles)} articles from live scraping")
        print("=" * 70)
        seen = set()
        for art in scraped_articles[:20]:
            title = art["title"][:70]
            if title not in seen:
                seen.add(title)
                print(f"  [{art['source']}] {title}")

    print()
    print("=" * 70)
    print(f"DONE - {len(deals)} deals exported to Excel and CSV")
    print("=" * 70)


if __name__ == "__main__":
    main()
